from fastmcp import FastMCP
import openpyxl
from memory_config import memory

# Excel yuklash
wb = openpyxl.load_workbook("transacts.xlsx")
ws = wb["Sheet1"]

# Ma'lumotlarni dict ga o'girish
headers = [cell.value for cell in ws[1]]
transactions = []
for row in ws.iter_rows(min_row=2, values_only=True):
    transactions.append(dict(zip(headers, row)))

# Personal info tools
personal = FastMCP("Personal")


def find_by(key: str, value: str):
    return [t for t in transactions 
            if value.lower() in str(t.get(key, "")).lower()]

@personal.tool()
def get_info_by_name(name: str) -> str:
    """Get card number, PINFL, phone number and date of birth by client full name"""
    results = find_by("CLIENT_FULL_NAME", name)
    if not results:
        return f"'{name}' ismli mijoz topilmadi"
    r = results[0]
    return (
        f"Ism: {r['CLIENT_FULL_NAME']}\n"
        f"PINFL: {r['PINFL']}\n"
        f"Karta: {r['CARD_ACCOUNT']}\n"
        f"Telefon: {r['PHONE_NUMBER']}\n"
        f"Tug'ilgan sana: {r['DATE_BIRTH']}"
    )

@personal.tool()
def get_info_by_pinfl(pinfl: str) -> str:
    """Get card number, name, phone number and date of birth by PINFL"""
    results = find_by("PINFL", pinfl)
    if not results:
        return f"PINFL {pinfl} topilmadi"
    r = results[0]
    return (
        f"Ism: {r['CLIENT_FULL_NAME']}\n"
        f"PINFL: {r['PINFL']}\n"
        f"Karta: {r['CARD_ACCOUNT']}\n"
        f"Telefon: {r['PHONE_NUMBER']}\n"
        f"Tug'ilgan sana: {r['DATE_BIRTH']}"
    )

@personal.tool()
def get_info_by_card(card_account: str) -> str:
    """Get name, PINFL, phone number and date of birth by card account number"""
    results = find_by("CARD_ACCOUNT", card_account)
    if not results:
        return f"Karta {card_account} topilmadi"
    r = results[0]
    return (
        f"Ism: {r['CLIENT_FULL_NAME']}\n"
        f"PINFL: {r['PINFL']}\n"
        f"Karta: {r['CARD_ACCOUNT']}\n"
        f"Telefon: {r['PHONE_NUMBER']}\n"
        f"Tug'ilgan sana: {r['DATE_BIRTH']}"
    )

@personal.tool()
def get_info_by_phone(phone_number: str) -> str:
    """Get name, PINFL, card number and date of birth by phone number"""
    results = find_by("PHONE_NUMBER", phone_number)
    if not results:
        return f"Telefon {phone_number} topilmadi"
    r = results[0]
    return (
        f"Ism: {r['CLIENT_FULL_NAME']}\n"
        f"PINFL: {r['PINFL']}\n"
        f"Karta: {r['CARD_ACCOUNT']}\n"
        f"Telefon: {r['PHONE_NUMBER']}\n"
        f"Tug'ilgan sana: {r['DATE_BIRTH']}"
    )
# Payroll server
payroll = FastMCP("Payroll")

@payroll.tool()
def get_transactions_by_pinfl(pinfl: str) -> str:
    """Get all transactions for a user by PINFL"""
    results = [t for t in transactions if str(t["PINFL"]) == pinfl]
    if not results:
        return f"PINFL {pinfl} uchun tranzaksiya topilmadi"
    output = f"Jami {len(results)} ta to'lov:\n"
    for t in results[:5]:  # faqat 5 tasi
        output += f"  {t['PAYMENT_DATE'].strftime('%d.%m.%Y')} - {t['AMOUNT']:,} so'm\n"
    return output

@payroll.tool()
def get_transaction_by_card(card_account: str) -> str:
    """Get transactions by card account number"""
    results = [t for t in transactions if str(t["CARD_ACCOUNT"]) == card_account]
    if not results:
        return f"Karta {card_account} uchun tranzaksiya topilmadi"
    output = f"Karta {card_account} uchun {len(results)} ta to'lov:\n"
    for t in results[:5]:
        output += f"  {t['CLIENT_FULL_NAME']} - {t['AMOUNT']:,} so'm\n"
    return output

@payroll.tool()
def get_total_by_period(period: str) -> str:
    """Get total payments for a specific period like 03.2026"""
    results = [t for t in transactions if t["PAYMENT_PERIOD"] == period]
    if not results:
        return f"{period} davri uchun ma'lumot topilmadi"
    total = sum(t["AMOUNT"] for t in results)
    return f"{period} davri: {len(results)} ta to'lov, jami {total:,} so'm"

# Mock Deposit server
deposit = FastMCP("Deposit")

@deposit.tool()
def get_balance(user_id: str) -> str:
    """Get deposit balance for user"""
    return f"User {user_id} deposit balance: 10,000,000 so'm"

@deposit.tool()
def get_history(user_id: str) -> str:
    """Get deposit transaction history"""
    return f"User {user_id} last transactions: +5M (2026-01), -2M (2026-02)"


# Mock Credit server
credit = FastMCP("Credit")

@credit.tool()
def get_limit(user_id: str) -> str:
    """Get credit limit for user"""
    return f"User {user_id} credit limit: 50,000,000 so'm"

@credit.tool()
def get_history(user_id: str) -> str:
    """Get credit payment history"""
    return f"User {user_id} paid: 2M (Jan), 2M (Feb), 2M (Mar)"


# Mock Card server
card = FastMCP("Card")

@card.tool()
def get_balance(user_id: str) -> str:
    """Get card balance for user"""
    return f"User {user_id} card balance: 2,500,000 so'm"

@card.tool()
def block_card(user_id: str) -> str:
    """Block card for user"""
    return f"User {user_id} card has been blocked successfully"

 




#Xotira toolari
mem0_server = FastMCP("Memory")

@mem0_server.tool()
def search_memory(query: str, user_id: str = "bank_user_001") -> str:
    """Search user memory for relevant information. Use this when you need to recall user preferences, personal info, or past interactions."""
    results = memory.search(query, user_id=user_id)
    if not results["results"]:
        return "Xotirada ma'lumot topilmadi"
    return "\n".join([f"- {r['memory']}" for r in results["results"]])

@mem0_server.tool()
def add_memory(text: str, user_id: str = "bank_user_001") -> str:
    """Save important information to user memory. Use this when user shares personal info, preferences, or important facts."""
    memory.add(text, user_id=user_id)
    return f"Xotiraga saqlandi: {text}"